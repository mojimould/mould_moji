# last update: 20240801
# Copyright 2023-2024 The individual creator, not held by any corporation.
# All rights reserved.

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from itertools import count
import xlwings as xw


# ファイルの読込み
with open('./subscripts/load_values.py', 'r', encoding='utf-8') as file:
    load_values = file.read()
exec(load_values)

# 関数
# 小数点以下第2位または第3位までを表示
def format_number(num):
    if num is None or type(num) not in [float, int]:
        return None

    str_num = "{:.3f}".format(num)
    if str_num[-1] == '0':
        str_num = "{:.2f}".format(num)
    return str_num

# テーパ#
WS2_name = taperID

# 余分なシートの削除
for sheet in WB.sheetnames:
    if sheet != WS.title and sheet != WS2_name:
        del WB[sheet]


# シートの作成
checkWS = WB.create_sheet(title=MainPrgID)

# 印刷時に水平方向を中央に設定
checkWS.page_setup.horizontalCentered = True

# 設定
file_name = MainPrgID + '.xlsx'
font_name = 'Meiryo UI'
font_size_title      = 17.0
font_size_yymmdd     = 12.0
font_size_drawingID  = 16.5
font_size_measurer   = 14.0
font_size_table_head = 14.0
font_size_item       = 14.0

preamble_col_height   = 26.0
last_row_col_height   = 34.0
table_col_height      = 25.0
table_head_col_height = 22.0

# footer text
# footer_text = MainPrgID + '：' + CustomerName

# table head
row_item = [
    "測定箇所",
    "狙い寸法",
    "公差範囲",
    "測定値1",
    "測定値2",
    "判定",
]

TitleName     = "横型MC 加工後 寸法・外観 確認表"
yymmddName    = "20       年       月       日"
DrawingIDName = "図面番号"
MeasurerName  = "測定者"

checkItem = "適 ・ 否"
JudgementItem = " 合 ・ 否 "

#
cornerRTolerance       = 0.5
dimpleDepthTolerance   = 0.05
dimpleOtherTolerance   = 0.5
keywayACDepthTolerance = 0.7
outcutLengthTolerance  = 0.5
otherTolerance         = 0.3

# 項目
totalLengthName                = "全長"
keywayPosName                  = "溝位置"
keywayWidthName                = "溝幅"
keywayACODName                 = "溝径AC"
keywayBDODName                 = "溝径BD"
if TopOutcutExistsFlag == 0 and CenterCurvatureExistsFlag == 1 or TopOutcutExistsFlag == 1 and TopCurvedOutcutExistsFlag == 1:
    if KeywayAsideDepthToleranceExitsFlag == 0:
        keywayAsideTopSide = "（トップ側）"
    else:
        keywayAsideTopSide = ""
    keywayCsideBottomSide = "（ボトム側）"
    keywayACDepthToleranceRange = '|A-C| ≦ ' + str(keywayACDepthTolerance)
elif TopOutcutExistsFlag == 1 and TopCurvedOutcutExistsFlag == 0:
    keywayAsideTopSide    = ""
    keywayCsideBottomSide = ""
    keywayACDepthToleranceRange = '|A-C| ≦ ' + str(otherTolerance)
AkeywayDepthName               = "溝深さA" + keywayAsideTopSide
CkeywayDepthName               = "溝深さC" + keywayCsideBottomSide
BkeywayDepthName               = "溝深さB"
DkeywayDepthName               = "溝深さD"
keywaCornerRName               = "溝コーナーR"
topOutcutACODName              = "トップ外削径AC"
topOutcutBDODName              = "トップ外削径BD"
topOutcutAsideThicknessName    = "トップ外削 A肉厚"
topOutcutCsideThicknessName    = "トップ外削 C肉厚"
topOutcutBsideThicknessName    = "トップ外削 B肉厚"
topOutcutDsideThicknessName    = "トップ外削 D肉厚"
topOutcutLengthName            = "トップ外削長"
topOutcutCornerRName           = "トップ外削コーナーR"
bottomOutcutACODName           = "ボトム外削径AC"
bottomOutcutBDODName           = "ボトム外削径BD"
bottomOutcutAsideThicknessName = "ボトム外削 A肉厚"
bottomOutcutCsideThicknessName = "ボトム外削 C肉厚"
bottomOutcutBsideThicknessName = "ボトム外削 B肉厚"
bottomOutcutDsideThicknessName = "ボトム外削 D肉厚"
bottomOutcutLengthName         = "ボトム外削長"
bottomOutcutCornerRName        = "ボトム外削コーナーR"
outcutACCenterDifName          = "外削 通り芯AC"
outcutBDCenterDifName          = "外削 通り芯BD"
endfaceBoringWidthName         = "端面座ぐり幅"
endfaceBoringDepthName         = "端面座ぐり深さ"
endfaceBoringLengthName        = "端面座ぐり長"
incutBoringACIDName            = "内削座ぐり径AC"
incutBoringBDIDName            = "内削座ぐり径BD"
incutBoringLengthName          = "内削座ぐり長"
incutBoringCornerRName         = "内削座ぐりコーナーR"
dimpleRowsName                 = "ディンプル列数"
dimpleNumOddRowName            = "1列目 ディンプル数"
dimpleNumEvenRowName           = "2列目 ディンプル数"
dimpleFirstRowDistanceName     = "ディンプル1列目 位置"
if DimpleExistsFlag == 1:
    if dimpleRowNumVal > 0:
        dimpleLastRowDistanceName  = "ディンプル" + f'{dimpleRowNumVal:.0f}' + "列目 位置"
dimpleRowsPitchName            = "ディンプル列間の間隔"
dimpleDepthName                = "ディンプル深さ"
appearanceDimpleName           = "外観 ディンプル全体"
BDcurvatureToleranceName       = "倒れBD / 外径"
ACcurvatureToleranceName       = "倒れAC / 外径"
appearanceEndfaceChamferName   = "外観 端面面取"
appearanceKeywayChamferName    = "外観 溝C面取"
appearanceEngravingName        = "外観 刻印"

col_item = []
# overall
col_item.append([
    totalLengthName,
    f'{totalLengthAve:.2f}',
    f'{totalLengthDim + totalLengthMinusTolerance:.2f} - ' + f'{totalLengthDim + totalLengthPlusTolerance:.2f}',
    '',
    '',
    JudgementItem
])

# keyway position
col_item.append([
    keywayPosName,
    format_number(keywayPosAve),
    format_number(keywayPosDim + keywayPosMinusTolerance) + ' - ' + format_number(keywayPosDim + keywayPosPlusTolerance),
    '',
    '',
    JudgementItem
])

# keyway width
col_item.append([
    keywayWidthName,
    format_number(keywayWidthAve),
    format_number(keywayWidthDim + keywayWidthMinusTolerance) + ' - ' + format_number(keywayWidthDim + keywayWidthPlusTolerance),
    '',
    '',
    JudgementItem
])

# keyway AC OD
if KeywayTypeFlag != 4: # not BD only
    col_item.append([
        keywayACODName,
        f'{keywayACODAve:.2f}',
        f'{keywayACODDim + keywayACODMinusTolerance:.2f} - ' + f'{keywayACODDim + keywayACODPlusTolerance:.2f}',
    '',
    '',
    JudgementItem
    ])

# keyway BD OD
col_item.append([
    keywayBDODName,
    f'{keywayBDODAve:.2f}',
    f'{keywayBDODDim + keywayBDODMinusTolerance:.2f} - ' + f'{keywayBDODDim + keywayBDODPlusTolerance:.2f}',
    '',
    '',
    JudgementItem
])

# keyway A-side depth
if KeywayAsideDepthToleranceExitsFlag == 0:
    col_item.append([
        AkeywayDepthName,
        f'({keywayAsideDepthDim:.2f})',
        keywayACDepthToleranceRange,
    '',
    '',
    JudgementItem
    ])
elif KeywayAsideDepthToleranceExitsFlag == 1:
    col_item.append([
        AkeywayDepthName,
        f'{keywayAsideDepthAve:.2f}',
        f'{keywayAsideDepthDim + keywayAsideDepthMinusTolerance:.2f} - ' + f'{keywayAsideDepthDim + keywayAsideDepthPlusTolerance:.2f}',
    '',
    '',
    JudgementItem
    ])

# keyway C-side depth
col_item.append([
    CkeywayDepthName,
    f'({keywayCsideDepthDim:.2f})',
    keywayACDepthToleranceRange,
    '',
    '',
    JudgementItem
])

# keyway B-side depth
col_item.append([
    BkeywayDepthName,
    f'({keywayBsideDepthDim:.2f})',
    '|B-D| ≦ ' + str(otherTolerance),
    '',
    '',
    JudgementItem
])

# keyway D-side depth
col_item.append([
    DkeywayDepthName,
    f'({keywayDsideDepthDim:.2f})',
    '|B-D| ≦ ' + str(otherTolerance),
    '',
    '',
    JudgementItem
])

# keyway corner R
if KeywayTypeFlag == 1: # corner R
    col_item.append([
        keywaCornerRName,
        f'{keywayCornerRDim:.1f}',
        f'{keywayCornerRDim - cornerRTolerance:.1f} - ' + f'{keywayCornerRDim + cornerRTolerance:.1f}',
        checkItem,
        checkItem,
        JudgementItem
    ])

# keyway corner C appearance
if KeywayTypeFlag == 2: # corner C
    col_item.append([
        appearanceKeywayChamferName,
        '-',
        '-',
        checkItem,
        checkItem,
        JudgementItem
    ])

# top outcut
if TopOutcutExistsFlag == 1:
    # top outcut AC OD
    col_item.append([
        topOutcutACODName,
        f'{topOutcutACODAve:.2f}',
        f'{topOutcutACODDim + topOutcutACODMinusTolerance:.2f} - ' + f'{topOutcutACODDim + topOutcutACODPlusTolerance:.2f}',
        '',
        '',
        JudgementItem
    ])

    # top outcut BD OD
    col_item.append([
        topOutcutBDODName,
        f'{topOutcutBDODAve:.2f}',
        f'{topOutcutBDODDim + topOutcutBDODMinusTolerance:.2f} - ' + f'{topOutcutBDODDim + topOutcutBDODPlusTolerance:.2f}',
        '',
        '',
        JudgementItem
    ])

    # top outcut A side thickness
    if BotOutcutExistsFlag == 0 or (BotOutcutExistsFlag == 1 and OutcutCenterlineBasementFlag == 2):
        if topOutcutAsideThicknessMinusTolerance == 0 and topOutcutAsideThicknessPlusTolerance == 0:
            col_item.append([
                topOutcutAsideThicknessName,
                f'{topOutcutAsideThicknessAve - mekkiThicknessDim:.2f}',
                '-',
                '',
                '',
                JudgementItem
            ])
        else:
            col_item.append([
                topOutcutAsideThicknessName,
                f'{topOutcutAsideThicknessAve - mekkiThicknessDim:.2f}',
                f'{topOutcutAsideThicknessDim - mekkiThicknessDim + topOutcutAsideThicknessMinusTolerance:.2f} - ' + f'{topOutcutAsideThicknessDim - mekkiThicknessDim + topOutcutAsideThicknessPlusTolerance:.2f}',
                '',
                '',
                JudgementItem
            ])

    # top outcut length
    if TopOutcutLengthNEKeywayFlag == 1:
        if topOutcutLengthMinusTolerance == 0 and topOutcutLengthPlusTolerance == 0:
            col_item.append([
                topOutcutLengthName,
                f'{topOutcutLengthAve:.2f}',
                f'{topOutcutLengthDim - outcutLengthTolerance:.2f} - ' + f'{topOutcutLengthDim + outcutLengthTolerance:.2f}',
                '',
                '',
                JudgementItem
            ])
        else:
            col_item.append([
                topOutcutLengthName,
                f'{topOutcutLengthAve:.2f}',
                f'{topOutcutLengthDim + topOutcutLengthMinusTolerance:.2f} - ' + f'{topOutcutLengthDim + topOutcutLengthPlusTolerance:.2f}',
                '',
                '',
                JudgementItem
            ])

    # top outcut corner R
    col_item.append([
        topOutcutCornerRName,
        f'{topOutcutCornerRDim:.1f}',
        f'{topOutcutCornerRDim - cornerRTolerance:.1f} - ' + f'{topOutcutCornerRDim + cornerRTolerance:.1f}',
        checkItem,
        checkItem,
        JudgementItem
    ])

# bot outcut
if BotOutcutExistsFlag == 1:
    # bot outcut AC OD
    col_item.append([
        bottomOutcutACODName,
        f'{botOutcutACODAve:.2f}',
        f'{botOutcutACODDim + botOutcutACODMinusTolerance:.2f} - ' + f'{botOutcutACODDim + botOutcutACODPlusTolerance:.2f}',
        '',
        '',
        JudgementItem
    ])

    # bot outcut BD OD
    col_item.append([
        bottomOutcutBDODName,
        f'{botOutcutBDODAve:.2f}',
        f'{botOutcutBDODDim + botOutcutBDODMinusTolerance:.2f} - ' + f'{botOutcutBDODDim + botOutcutBDODPlusTolerance:.2f}',
        '',
        '',
        JudgementItem
    ])

    # bot outcut A side thickness
    if TopOutcutExistsFlag == 0 or (TopOutcutExistsFlag == 1 and OutcutCenterlineBasementFlag == 1):
        if botOutcutAsideThicknessMinusTolerance == 0 and botOutcutAsideThicknessPlusTolerance == 0:
            col_item.append([
                bottomOutcutAsideThicknessName,
                f'{botOutcutAsideThicknessAve - mekkiThicknessDim:.2f}',
                '-',
                '',
                '',
                JudgementItem
            ])
        else:
            col_item.append([
                bottomOutcutAsideThicknessName,
                f'{botOutcutAsideThicknessAve - mekkiThicknessDim:.2f}',
                f'{botOutcutAsideThicknessDim - mekkiThicknessDim + botOutcutAsideThicknessMinusTolerance:.2f} - ' + f'{botOutcutAsideThicknessDim - mekkiThicknessDim + botOutcutAsideThicknessPlusTolerance:.2f}',
                '',
                '',
                JudgementItem
            ])
    else:
        col_item.append([
            bottomOutcutAsideThicknessName,
            '-',
            '-',
            '',
            '',
            JudgementItem
        ])

    # bot outcut length
    if botOutcutLengthMinusTolerance == 0 and botOutcutLengthPlusTolerance == 0:
        col_item.append([
            bottomOutcutLengthName,
            f'{botOutcutLengthAve:.2f}',
            f'{botOutcutLengthDim - outcutLengthTolerance:.2f} - ' + f'{botOutcutLengthDim + outcutLengthTolerance:.2f}',
            '',
            '',
            JudgementItem
        ])
    else:
        col_item.append([
            bottomOutcutLengthName,
            f'{botOutcutLengthAve:.2f}',
            f'{botOutcutLengthDim + botOutcutLengthMinusTolerance:.2f} - ' + f'{botOutcutLengthDim + botOutcutLengthPlusTolerance:.2f}',
            '',
            '',
            JudgementItem
        ])

    # bot outcut corner R
    col_item.append([
        bottomOutcutCornerRName,
        f'{botOutcutCornerRDim:.1f}',
        f'{botOutcutCornerRDim - cornerRTolerance:.1f} - ' + f'{botOutcutCornerRDim + cornerRTolerance:.1f}',
        checkItem,
        checkItem,
        JudgementItem
    ])

# centerline difference
if TopOutcutExistsFlag == 1 and BotOutcutExistsFlag == 1:
    if centerlineACDifMinusTolerance == 0 and centerlineACDifPlusTolerance == 0:
        centerlineACDifMinusTolerance = -0.5
        centerlineACDifPlusTolerance = 0.5
    col_item.append([
        outcutACCenterDifName,
        f'{centerlineACDifAve:.2f}',
        f'{centerlineACDifDim + centerlineACDifMinusTolerance:.2f} - ' + f'{centerlineACDifDim + centerlineACDifPlusTolerance:.2f}',
        '',
        '',
        JudgementItem
    ])
    centerlineBDDifMinusTolerance = -0.5
    centerlineBDDifPlusTolerance = 0.5
    col_item.append([
        outcutBDCenterDifName,
        0,
        f'{0 + centerlineBDDifMinusTolerance:.2f} - ' + f'{0 + centerlineBDDifPlusTolerance:.2f}',
        '',
        '',
        JudgementItem
    ])

# dimple row num
if DimpleExistsFlag == 1:
    col_item.append([
        dimpleRowsName,
        f'{dimpleRowNumVal:.0f}',
        f'{dimpleRowNumVal:.0f}',
        '',
        '',
        JudgementItem
    ])

    # dimple num odd row
    col_item.append([
        dimpleNumOddRowName,
        f'{dimpleFirstRowLengthDim / dimpleHorizontalPitchDim + 1:.0f}',
        f'{dimpleFirstRowLengthDim / dimpleHorizontalPitchDim + 1:.0f}',
        '',
        '',
        JudgementItem
    ])

    # dimple num even row
    col_item.append([
        dimpleNumEvenRowName,
        f'{dimpleSecondRowLengthDim / dimpleHorizontalPitchDim + 1:.0f}',
        f'{dimpleSecondRowLengthDim / dimpleHorizontalPitchDim + 1:.0f}',
        '',
        '',
        JudgementItem
    ])

    # dimple 1st row distance from end face
    col_item.append([
        dimpleFirstRowDistanceName,
        f'{dimpleFirstRowDistanceFromEndfaceDim:.1f}',
        f'{dimpleFirstRowDistanceFromEndfaceDim - dimpleOtherTolerance:.1f} - ' + f'{dimpleFirstRowDistanceFromEndfaceDim + dimpleOtherTolerance:.1f}',
        checkItem,
        checkItem,
        JudgementItem
    ])

    # dimple last row distance from end face
    col_item.append([
        dimpleLastRowDistanceName,
        f'{dimpleLastRowDistanceFromEndfaceDim:.1f}',
        f'{dimpleLastRowDistanceFromEndfaceDim - dimpleOtherTolerance:.1f} - ' + f'{dimpleLastRowDistanceFromEndfaceDim + dimpleOtherTolerance:.1f}',
        checkItem,
        checkItem,
        JudgementItem
    ])

    # dimple depth
    col_item.append([
        dimpleDepthName,
        f'{dimpleDepthDim:.2f}',
        f'{dimpleDepthDim - dimpleDepthTolerance:.2f} - ' + f'{dimpleDepthDim + dimpleDepthTolerance:.2f}',
        checkItem,
        checkItem,
        JudgementItem
    ])

    # dimple appearance
    col_item.append([
        appearanceDimpleName,
        '-',
        '-',
        checkItem,
        checkItem,
        JudgementItem
    ])

# BD curvature Tolerance
col_item.append([
    BDcurvatureToleranceName,
    '0',
    '0.0 - 0.1',
    checkItem,
    checkItem,
    JudgementItem
])

# AC curvature Tolerance
col_item.append([
    ACcurvatureToleranceName,
    '0',
    '0.0 - 0.1',
    checkItem,
    checkItem,
    JudgementItem
])

# endface chamfer appearance
col_item.append([
    appearanceEndfaceChamferName,
    '-',
    '-',
    checkItem,
    checkItem,
    JudgementItem
])

# engraving appearance
col_item.append([
    appearanceEngravingName,
    '-',
    '-',
    checkItem,
    checkItem,
    JudgementItem
])


title_font      = Font(bold=True, size=font_size_title, name=font_name)
yymmdd_font     = Font(size=font_size_yymmdd, name=font_name)
drawingID_font  = Font(size=font_size_drawingID, name=font_name)
measurer_font   = Font(size=font_size_measurer, name=font_name)
table_haed_font = Font(bold=True, size=font_size_table_head, name=font_name)
table_item_font = Font(size=font_size_item, name=font_name)

# 罫線のスタイルを定義
thin_border = Border(
    left   = Side(style='thin'), 
    right  = Side(style='thin'), 
    top    = Side(style='thin'), 
    bottom = Side(style='thin')
)
top_double_border = Border(
    left   = Side(style='thin'), 
    right  = Side(style='thin'), 
    top    = Side(style='double'), 
    bottom = Side(style='thin')
)
thin_border_side = Side(style='thin')
no_border_side = Side(style=None)
# 二重線の項目
doubleBorderItem = [
    totalLengthName,
    keywayPosName,
    topOutcutACODName,
    bottomOutcutACODName,
    outcutACCenterDifName,
    endfaceBoringWidthName,
    incutBoringACIDName,
    dimpleRowsName,
    BDcurvatureToleranceName,
    appearanceEndfaceChamferName
]

# 項目を追加
row_num = 4
for i, item in enumerate(row_item, start=1):
    itemCell = checkWS.cell(row=row_num, column=i, value=item)
    itemCell.font = table_haed_font
    itemCell.border = thin_border  # 罫線を追加
    # 行の高さを設定
    checkWS.row_dimensions[row_num].height = table_head_col_height

row_num = row_num + 1
for row_item in col_item:
    for i, item in enumerate(row_item, start=1):
        itemCell = checkWS.cell(row=row_num, column=i, value=item)
        itemCell.font = table_item_font
        if checkWS.cell(row=row_num, column=1).value in doubleBorderItem: # 罫線を追加
            itemCell.border = top_double_border
        else:
            itemCell.border = thin_border
    # 行の高さを設定
    checkWS.row_dimensions[row_num].height = table_col_height
    row_num += 1 

# A列を左寄せに設定
for row in checkWS['A']:
    row.alignment = Alignment(horizontal='left', vertical='center')

# B-F列を中央揃えに設定
for col in ['B', 'C', 'D', 'E', 'F']:
    for cell in checkWS[col]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# ファイルを保存
WB.save(file_name)


# xlwingsでセル幅を自動調整
app = xw.App(visible=False)
wb_xlw = xw.Book(file_name)
sheet = wb_xlw.sheets[-1]
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    sheet.range(f'{col}1:{col}1048576').api.EntireColumn.AutoFit()

# セルの結合
sheet.range('B3:C3').merge()

# セル幅を取得
mergeCellWidth = sum([sheet.range(col + '3').column_width for col in ['B', 'C']]) + 1

# 結合したセルにテキストを入力
mergeCell = sheet['B3']
mergeCell.value = CustomerName
mergeCell.api.Font.Size = font_size_drawingID

# 縮小して全体を表示する設定
mergeCell.api.ShrinkToFit = True

# フォントサイズの取得
mergeCellFontSize = mergeCell.api.Font.Size
while True:
    tempCell = sheet['H3']
    tempCell.value = CustomerName
    tempCell.api.Font.Size = mergeCellFontSize
    sheet.range('H:H').api.EntireColumn.AutoFit()
    tempWidth = sheet.range('H3').column_width
    if tempWidth <= mergeCellWidth:
        break
    mergeCellFontSize -= 0.5
    mergeCell.api.Font.Size = mergeCellFontSize
sheet.range('H:H').delete()
mergeCell.value = ""

# 縮小して全体を表示する設定の解除
mergeCell.api.ShrinkToFit = False

# セルの結合を解除
sheet.range('B3:C3').unmerge()

# テキストを再度入力（結合解除後のセルに）
mergeCell.value = CustomerName
mergeCell.api.Font.Size = mergeCellFontSize
mergeCell.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignLeft

wb_xlw.save()
wb_xlw.close()



# openpyxl で再度読み込み
WB = load_workbook(file_name, data_only=True)
# 余分なシートの削除
for sheet in WB.sheetnames:
    if sheet != str(DrawingIDNum).zfill(4):
        del WB[sheet]

checkWS = WB.worksheets[0]

last_row = checkWS.max_row + 1
checkWS['F' + str(last_row)] = '振分：トップ ' + str(topAlocationLengthAve) + ', ボトム ' + str(botAlocationLengthAve)
checkWS['F' + str(last_row)].alignment = Alignment(horizontal='right', vertical='top')
checkWS.row_dimensions[last_row].height = last_row_col_height

# 最終行のAからFまでのセルに罫線を適用
for i, cell in enumerate(checkWS['A' + str(checkWS.max_row):'F' + str(checkWS.max_row)][0]):
    border_left = thin_border_side if i == 0 else no_border_side
    border_right = thin_border_side if i == 5 else no_border_side
    cell.border = Border(left=border_left, right=border_right, top=thin_border_side, bottom=thin_border_side)

checkWS['A1'].value = TitleName
checkWS['A1'].font = title_font
checkWS['A1'].alignment = Alignment(horizontal='left', vertical='top')
checkWS.row_dimensions[1].height = preamble_col_height

checkWS['A2'].value = yymmddName
checkWS['A2'].font = yymmdd_font
checkWS['A2'].alignment = Alignment(horizontal='left', vertical='center')
checkWS.row_dimensions[2].height = preamble_col_height

checkWS['A3'].value = DrawingIDName + '： ' + MainPrgID
checkWS['A3'].font = drawingID_font
checkWS['A3'].alignment = Alignment(horizontal='left', vertical='top')
checkWS.row_dimensions[3].height = preamble_col_height

checkWS['D2'].value = MeasurerName + '1'
checkWS['D2'].font = measurer_font
checkWS['D2'].alignment = Alignment(horizontal='center', vertical='bottom')
checkWS['D3'].border = thin_border

checkWS['E2'].value = MeasurerName + '2'
checkWS['E2'].font = measurer_font
checkWS['E2'].alignment = Alignment(horizontal='center', vertical='bottom')
checkWS['E3'].border = thin_border

# 最後の行と最後の列を取得
last_row = checkWS.max_row
last_column = checkWS.max_column

# 最後の列のアルファベットラベルを取得
last_column_letter = get_column_letter(last_column)

# 用紙サイズ
checkWS.page_setup.paperSize = checkWS.PAPERSIZE_A4

# 向き
checkWS.page_setup.orientation = "portrait"

# 印刷範囲
checkWS.page_setup.print_area = f'A1:{last_column_letter}{last_row}'

# 印刷範囲を1ページに収める
checkWS.page_setup.fitToPage = True
checkWS.page_setup.fitToHeight = 1
checkWS.page_setup.fitToWidth = 1

# 余白（インチ）
checkWS.page_margins.left = 0.75
checkWS.page_margins.right = 0.1
checkWS.page_margins.top = 0.0
checkWS.page_margins.bottom = 0.3
checkWS.page_margins.header = 0.0
checkWS.page_margins.footer = 0.0

# フッターのテキスト
# checkWS.oddFooter.right.text = footer_text

# 印刷時に水平方向を中央に設定
checkWS.print_options.horizontalCentered = True

# ファイルを保存
WB.save(file_name)
